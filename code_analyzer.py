import os
import re
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from datetime import datetime

# ─── CONSTANTES DE SEGURIDAD ────────────────────────────────────
MAX_RUTA_LENGTH = 500
MAX_ARCHIVOS = 5000
MAX_LINEAS_ARCHIVO = 100000
EXTENSIONES_PERMITIDAS = {".java", ".cs"}
CARPETAS_IGNORADAS = {
    ".git", "target", "bin", "obj", ".idea",
    "node_modules", ".mvn", "out", "build",
    "__pycache__", ".vs", "dist", "release"
}

# ─── ESTILOS ────────────────────────────────────────────────────
DARK_BLUE  = "1A3A5C"
MID_BLUE   = "2E6DA4"
LIGHT_BLUE = "D6E4F0"
WHITE      = "FFFFFF"
DARK_GRAY  = "343A40"

def header_fill(color):
    return PatternFill("solid", fgColor=color)

def header_font(color=WHITE, bold=True, size=11):
    return Font(name="Calibri", color=color, bold=bold, size=size)

def data_font(bold=False, size=10, color=DARK_GRAY):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

# ─── VALIDACIONES ───────────────────────────────────────────────

def validar_ruta(ruta):
    """Valida que la ruta sea segura y exista."""
    if not ruta or not isinstance(ruta, str):
        return False, "La ruta no puede estar vacía."

    if len(ruta) > MAX_RUTA_LENGTH:
        return False, f"La ruta no puede superar {MAX_RUTA_LENGTH} caracteres."

    # Resolver ruta absoluta y verificar que existe
    try:
        ruta_abs = Path(ruta).resolve()
    except Exception:
        return False, "La ruta contiene caracteres inválidos."

    if not ruta_abs.exists():
        return False, f"La ruta '{ruta}' no existe."

    if not ruta_abs.is_dir():
        return False, f"La ruta '{ruta}' no es una carpeta."

    return True, str(ruta_abs)

def validar_lenguaje(opcion):
    """Valida la opción de lenguaje seleccionada."""
    if opcion not in ("1", "2"):
        return False, "Opción inválida. Selecciona 1 (Java) o 2 (C#)."
    return True, "java" if opcion == "1" else "csharp"

def es_archivo_seguro(ruta_archivo):
    """Verifica que el archivo sea seguro para leer."""
    try:
        ruta = Path(ruta_archivo)

        # Verificar extensión permitida
        if ruta.suffix.lower() not in EXTENSIONES_PERMITIDAS:
            return False

        # Verificar que el archivo no sea un enlace simbólico
        if ruta.is_symlink():
            return False

        # Verificar tamaño máximo (5 MB)
        if ruta.stat().st_size > 5 * 1024 * 1024:
            return False

        return True
    except Exception:
        return False

def sanitizar_texto(texto):
    """Elimina caracteres potencialmente peligrosos para Excel."""
    if not texto or not isinstance(texto, str):
        return ""
    # Eliminar caracteres de control excepto saltos de línea
    texto = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', texto)
    # Limitar longitud
    return texto[:500]

# ─── ANÁLISIS DE CÓDIGO ─────────────────────────────────────────

def analizar_java(contenido):
    """Analiza un archivo Java de forma segura."""
    try:
        lineas = contenido.split("\n")

        # Limitar análisis si el archivo es demasiado grande
        if len(lineas) > MAX_LINEAS_ARCHIVO:
            return None, f"Archivo demasiado grande ({len(lineas)} líneas)."

        lineas_totales = len(lineas)
        lineas_codigo = sum(
            1 for l in lineas
            if l.strip()
            and not l.strip().startswith("//")
            and not l.strip().startswith("*")
            and not l.strip().startswith("/*")
        )

        patron_clase = re.compile(
            r'(?:public|private|protected)?\s*(?:abstract|final)?\s*'
            r'(?:class|interface|enum)\s+(\w+)'
        )
        nombres_clases = patron_clase.findall(contenido)

        patron_metodo = re.compile(
            r'(?:public|private|protected|static|\s)+[\w\<\>\[\]]+\s+'
            r'(\w+)\s*\([^)]*\)\s*(?:throws\s+[\w,\s]+)?\s*\{'
        )
        palabras_reservadas = {
            "if", "while", "for", "switch", "catch",
            "try", "else", "finally", "do", "synchronized"
        }
        nombres_metodos = [
            m for m in patron_metodo.findall(contenido)
            if m not in palabras_reservadas
        ]

        patron_paquete = re.compile(r'^package\s+([\w.]+);', re.MULTILINE)
        match_paquete = patron_paquete.search(contenido)
        paquete = match_paquete.group(1) if match_paquete else "sin paquete"

        return {
            "paquete": sanitizar_texto(paquete),
            "clases": [sanitizar_texto(c) for c in nombres_clases],
            "metodos": [sanitizar_texto(m) for m in nombres_metodos],
            "lineas_totales": lineas_totales,
            "lineas_codigo": lineas_codigo,
        }, None
    except re.error as e:
        return None, f"Error en expresión regular: {e}"
    except Exception as e:
        return None, f"Error inesperado: {e}"


def analizar_csharp(contenido):
    """Analiza un archivo C# de forma segura."""
    try:
        lineas = contenido.split("\n")

        if len(lineas) > MAX_LINEAS_ARCHIVO:
            return None, f"Archivo demasiado grande ({len(lineas)} líneas)."

        lineas_totales = len(lineas)
        lineas_codigo = sum(
            1 for l in lineas
            if l.strip()
            and not l.strip().startswith("//")
            and not l.strip().startswith("*")
            and not l.strip().startswith("/*")
        )

        patron_clase = re.compile(
            r'(?:public|private|protected|internal)?\s*'
            r'(?:abstract|sealed|static)?\s*'
            r'(?:class|interface|enum|struct)\s+(\w+)'
        )
        nombres_clases = patron_clase.findall(contenido)

        patron_metodo = re.compile(
            r'(?:public|private|protected|internal|static|virtual|override|async|\s)+'
            r'[\w\<\>\[\]]+\s+(\w+)\s*\([^)]*\)\s*(?:\{|=>)'
        )
        palabras_reservadas = {
            "if", "while", "for", "switch", "catch",
            "try", "else", "finally", "do", "lock"
        }
        nombres_metodos = [
            m for m in patron_metodo.findall(contenido)
            if m not in palabras_reservadas
        ]

        patron_ns = re.compile(r'^namespace\s+([\w.]+)', re.MULTILINE)
        match_ns = patron_ns.search(contenido)
        namespace = match_ns.group(1) if match_ns else "sin namespace"

        return {
            "paquete": sanitizar_texto(namespace),
            "clases": [sanitizar_texto(c) for c in nombres_clases],
            "metodos": [sanitizar_texto(m) for m in nombres_metodos],
            "lineas_totales": lineas_totales,
            "lineas_codigo": lineas_codigo,
        }, None
    except re.error as e:
        return None, f"Error en expresión regular: {e}"
    except Exception as e:
        return None, f"Error inesperado: {e}"


def escanear_proyecto(ruta, lenguaje):
    """Recorre la carpeta del proyecto de forma segura."""
    extension = ".java" if lenguaje == "java" else ".cs"
    resultados = []
    errores = []
    archivos_procesados = 0

    ruta_base = Path(ruta).resolve()

    for root, dirs, files in os.walk(ruta_base):
        # Ignorar carpetas no permitidas
        dirs[:] = [
            d for d in dirs
            if d not in CARPETAS_IGNORADAS
            and not d.startswith(".")
        ]

        for archivo in files:
            if archivos_procesados >= MAX_ARCHIVOS:
                errores.append(f"Límite de {MAX_ARCHIVOS} archivos alcanzado.")
                return resultados, errores

            if not archivo.endswith(extension):
                continue

            ruta_completa = Path(root) / archivo

            # Verificar que el archivo está dentro de la ruta base
            try:
                ruta_completa.resolve().relative_to(ruta_base)
            except ValueError:
                errores.append(f"Archivo fuera de la ruta base ignorado: {archivo}")
                continue

            if not es_archivo_seguro(ruta_completa):
                errores.append(f"Archivo ignorado por seguridad: {archivo}")
                continue

            try:
                with open(ruta_completa, "r", encoding="utf-8", errors="ignore") as f:
                    contenido = f.read()

                if lenguaje == "java":
                    analisis, error = analizar_java(contenido)
                else:
                    analisis, error = analizar_csharp(contenido)

                if error:
                    errores.append(f"{archivo}: {error}")
                    continue

                ruta_relativa = str(ruta_completa.relative_to(ruta_base))
                analisis["archivo"] = sanitizar_texto(archivo)
                analisis["ruta"] = sanitizar_texto(ruta_relativa)
                resultados.append(analisis)
                archivos_procesados += 1
                print(f"  ✓ Analizado: {ruta_relativa}")

            except PermissionError:
                errores.append(f"Sin permisos para leer: {archivo}")
            except Exception as e:
                errores.append(f"Error en {archivo}: {sanitizar_texto(str(e))}")

    return resultados, errores


# ─── GENERACIÓN DEL EXCEL ───────────────────────────────────────

def validar_ruta_salida(ruta_salida):
    """Valida que la ruta de salida sea segura."""
    try:
        ruta = Path(ruta_salida).resolve()

        # Verificar que el directorio padre existe
        if not ruta.parent.exists():
            return False, "El directorio de salida no existe."

        # Verificar que tiene extensión xlsx
        if ruta.suffix.lower() != ".xlsx":
            return False, "El archivo de salida debe tener extensión .xlsx"

        # Verificar que el nombre no tiene caracteres peligrosos
        nombre = ruta.name
        if re.search(r'[<>:"/\\|?*]', nombre):
            return False, "El nombre del archivo contiene caracteres inválidos."

        return True, str(ruta)
    except Exception as e:
        return False, f"Ruta de salida inválida: {e}"


def aplicar_estilo_header(ws, fila, num_cols, color=DARK_BLUE):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=fila, column=col)
        cell.fill = header_fill(color)
        cell.font = header_font()
        cell.alignment = center()
        cell.border = thin_border()


def generar_excel(resultados, ruta_salida, nombre_proyecto, lenguaje):
    """Genera el reporte Excel de forma segura."""
    # Validar ruta de salida
    valido, ruta_validada = validar_ruta_salida(ruta_salida)
    if not valido:
        raise ValueError(f"Ruta de salida inválida: {ruta_validada}")

    wb = Workbook()

    # ── Hoja 1: Resumen ─────────────────────────────────────────
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    ws_resumen.column_dimensions["A"].width = 45
    ws_resumen.column_dimensions["B"].width = 20

    ws_resumen.merge_cells("A1:B1")
    t = ws_resumen["A1"]
    t.value = "CodeAnalyzer — Reporte de Proyecto"
    t.fill = header_fill(DARK_BLUE)
    t.font = Font(name="Calibri", color=WHITE, bold=True, size=14)
    t.alignment = center()
    ws_resumen.row_dimensions[1].height = 30

    ws_resumen.merge_cells("A2:B2")
    t2 = ws_resumen["A2"]
    nombre_seguro = sanitizar_texto(nombre_proyecto)
    t2.value = (f"Proyecto: {nombre_seguro}  |  "
                f"Lenguaje: {lenguaje.upper()}  |  "
                f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    t2.fill = header_fill(MID_BLUE)
    t2.font = Font(name="Calibri", color=WHITE, size=10)
    t2.alignment = center()

    total_archivos = len(resultados)
    total_clases   = sum(len(r["clases"]) for r in resultados)
    total_metodos  = sum(len(r["metodos"]) for r in resultados)
    total_lineas   = sum(r["lineas_totales"] for r in resultados)
    total_codigo   = sum(r["lineas_codigo"] for r in resultados)

    metricas = [
        ("Total de archivos analizados", total_archivos),
        ("Total de clases encontradas", total_clases),
        ("Total de métodos encontrados", total_metodos),
        ("Total de líneas (con vacías y comentarios)", total_lineas),
        ("Total de líneas de código puro", total_codigo),
        ("Promedio de líneas por archivo",
         round(total_lineas / total_archivos, 1) if total_archivos else 0),
        ("Promedio de métodos por clase",
         round(total_metodos / total_clases, 1) if total_clases else 0),
    ]

    for i, (etiqueta, valor) in enumerate(metricas):
        fila = 4 + i
        ws_resumen.cell(row=fila, column=1, value=etiqueta)
        ws_resumen.cell(row=fila, column=2, value=valor)
        alternado = i % 2 == 0
        for col in [1, 2]:
            cell = ws_resumen.cell(row=fila, column=col)
            cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE if alternado else WHITE)
            cell.font = data_font(bold=(col == 1))
            cell.alignment = left() if col == 1 else center()
            cell.border = thin_border()

    # ── Hoja 2: Detalle por Archivo ──────────────────────────────
    ws_detalle = wb.create_sheet("Detalle por Archivo")
    anchos = [30, 50, 25, 12, 12, 18, 18]
    letras = ["A", "B", "C", "D", "E", "F", "G"]
    for letra, ancho in zip(letras, anchos):
        ws_detalle.column_dimensions[letra].width = ancho

    ws_detalle.merge_cells("A1:G1")
    t = ws_detalle["A1"]
    t.value = "Detalle por Archivo"
    t.fill = header_fill(DARK_BLUE)
    t.font = Font(name="Calibri", color=WHITE, bold=True, size=13)
    t.alignment = center()

    headers = ["Archivo", "Ruta", "Paquete / Namespace",
               "Clases", "Métodos", "Líneas Totales", "Líneas de Código"]
    for col, h in enumerate(headers, 1):
        cell = ws_detalle.cell(row=2, column=col, value=h)
        cell.fill = header_fill(MID_BLUE)
        cell.font = header_font()
        cell.alignment = center()
        cell.border = thin_border()

    for i, r in enumerate(resultados):
        fila = 3 + i
        valores = [
            r["archivo"], r["ruta"], r["paquete"],
            len(r["clases"]), len(r["metodos"]),
            r["lineas_totales"], r["lineas_codigo"],
        ]
        for col, val in enumerate(valores, 1):
            cell = ws_detalle.cell(row=fila, column=col, value=val)
            cell.fill = PatternFill("solid",
                                    fgColor=LIGHT_BLUE if i % 2 == 0 else WHITE)
            cell.font = data_font()
            cell.alignment = left() if col <= 3 else center()
            cell.border = thin_border()

    # ── Hoja 3: Métodos ──────────────────────────────────────────
    ws_metodos = wb.create_sheet("Métodos")
    ws_metodos.column_dimensions["A"].width = 35
    ws_metodos.column_dimensions["B"].width = 35
    ws_metodos.column_dimensions["C"].width = 40

    ws_metodos.merge_cells("A1:C1")
    t = ws_metodos["A1"]
    t.value = "Listado Completo de Métodos"
    t.fill = header_fill(DARK_BLUE)
    t.font = Font(name="Calibri", color=WHITE, bold=True, size=13)
    t.alignment = center()

    for col, h in enumerate(["Método", "Clase(s) en el Archivo", "Archivo"], 1):
        cell = ws_metodos.cell(row=2, column=col, value=h)
        cell.fill = header_fill(MID_BLUE)
        cell.font = header_font()
        cell.alignment = center()
        cell.border = thin_border()

    fila_m = 3
    for r in resultados:
        clases_str = ", ".join(r["clases"]) if r["clases"] else "sin clase"
        for metodo in r["metodos"]:
            valores = [metodo, clases_str[:200], r["archivo"]]
            for col, val in enumerate(valores, 1):
                cell = ws_metodos.cell(row=fila_m, column=col, value=val)
                cell.fill = PatternFill("solid",
                                        fgColor=LIGHT_BLUE if fila_m % 2 == 0 else WHITE)
                cell.font = data_font()
                cell.alignment = left()
                cell.border = thin_border()
            fila_m += 1

    # ── Hoja 4: Gráfico ──────────────────────────────────────────
    ws_grafico = wb.create_sheet("Gráfico")
    ws_grafico.column_dimensions["A"].width = 35
    ws_grafico.column_dimensions["B"].width = 20

    ws_grafico.merge_cells("A1:B1")
    t = ws_grafico["A1"]
    t.value = "Top 10 Archivos con más Líneas de Código"
    t.fill = header_fill(DARK_BLUE)
    t.font = Font(name="Calibri", color=WHITE, bold=True, size=13)
    t.alignment = center()

    for col, h in enumerate(["Archivo", "Líneas de Código"], 1):
        cell = ws_grafico.cell(row=2, column=col, value=h)
        cell.fill = header_fill(MID_BLUE)
        cell.font = header_font()
        cell.alignment = center()
        cell.border = thin_border()

    top10 = sorted(resultados,
                   key=lambda x: x["lineas_codigo"], reverse=True)[:10]
    for i, r in enumerate(top10):
        fila = 3 + i
        ws_grafico.cell(row=fila, column=1, value=r["archivo"])
        ws_grafico.cell(row=fila, column=2, value=r["lineas_codigo"])
        for col in [1, 2]:
            cell = ws_grafico.cell(row=fila, column=col)
            cell.fill = PatternFill("solid",
                                    fgColor=LIGHT_BLUE if i % 2 == 0 else WHITE)
            cell.font = data_font()
            cell.alignment = left() if col == 1 else center()
            cell.border = thin_border()

    chart = BarChart()
    chart.type = "col"
    chart.title = "Top 10 Archivos por Líneas de Código"
    chart.y_axis.title = "Líneas de Código"
    chart.x_axis.title = "Archivo"
    chart.style = 10
    chart.width = 20
    chart.height = 12

    data_ref = Reference(ws_grafico, min_col=2, min_row=2,
                         max_row=2 + len(top10))
    cats_ref = Reference(ws_grafico, min_col=1, min_row=3,
                         max_row=2 + len(top10))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws_grafico.add_chart(chart, "D2")

    # Guardar de forma segura
    wb.save(ruta_validada)
    print(f"\n✅ Reporte generado: {ruta_validada}")


# ─── PUNTO DE ENTRADA ───────────────────────────────────────────

def main():
    print("=" * 55)
    print("       CodeAnalyzer — Analizador de Código Fuente")
    print("       Versión 1.1.0 — Uso responsable únicamente")
    print("=" * 55)

    # Solicitar y validar ruta
    ruta_input = input("\n📁 Ruta del proyecto (Enter para carpeta actual): ").strip()
    if not ruta_input:
        ruta_input = os.getcwd()

    valido, resultado = validar_ruta(ruta_input)
    if not valido:
        print(f"\n❌ {resultado}")
        sys.exit(1)

    ruta = resultado

    # Solicitar y validar lenguaje
    print("\n🔤 Lenguaje:")
    print("   1. Java")
    print("   2. C#")
    opcion = input("Selecciona (1 o 2): ").strip()

    valido, lenguaje = validar_lenguaje(opcion)
    if not valido:
        print(f"\n❌ {lenguaje}")
        sys.exit(1)

    nombre_proyecto = sanitizar_texto(Path(ruta).name)
    print(f"\n🔍 Analizando proyecto: {nombre_proyecto}")
    print(f"   Lenguaje: {'Java' if lenguaje == 'java' else 'C#'}")
    print(f"   Ruta: {ruta}\n")

    # Escanear archivos
    resultados, errores = escanear_proyecto(ruta, lenguaje)

    # Mostrar errores no fatales
    if errores:
        print(f"\n⚠️  Advertencias durante el análisis ({len(errores)}):")
        for err in errores[:10]:
            print(f"   • {err}")
        if len(errores) > 10:
            print(f"   ... y {len(errores) - 10} advertencias más.")

    if not resultados:
        ext = ".java" if lenguaje == "java" else ".cs"
        print(f"\n❌ No se encontraron archivos {ext} en la ruta indicada.")
        sys.exit(1)

    print(f"\n📊 Archivos analizados: {len(resultados)}")

    # Generar Excel en la misma carpeta del script
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_seguro = re.sub(r'[^a-zA-Z0-9_-]', '_', nombre_proyecto)
    nombre_excel = f"Reporte_{nombre_seguro}_{timestamp}.xlsx"
    ruta_salida = str(Path(__file__).parent / nombre_excel)

    print("\n📝 Generando reporte Excel...")

    try:
        generar_excel(resultados, ruta_salida, nombre_proyecto, lenguaje)
    except ValueError as e:
        print(f"\n❌ Error de validación: {e}")
        sys.exit(1)
    except PermissionError:
        print("\n❌ Sin permisos para escribir en la carpeta destino.")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ Error al generar el reporte: {sanitizar_texto(str(e))}")
        sys.exit(1)

    print(f"\n📈 Resumen:")
    print(f"   Archivos : {len(resultados)}")
    print(f"   Clases   : {sum(len(r['clases']) for r in resultados)}")
    print(f"   Métodos  : {sum(len(r['metodos']) for r in resultados)}")
    print(f"   Líneas   : {sum(r['lineas_totales'] for r in resultados)}")
    print("\n✅ ¡Listo! Abre el archivo Excel para ver el reporte completo.")
    input("\nPresiona Enter para salir...")


if __name__ == "__main__":
    main()